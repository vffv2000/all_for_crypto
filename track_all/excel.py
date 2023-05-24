import openpyxl


def readXLSX():
    # Открываем файл Excel
    workbook = openpyxl.load_workbook('balance.xlsx')

    # Получаем активный лист
    sheet = workbook.active

    # Читаем адреса из ячеек B5 и ниже
    addresses = []
    for cell in sheet['B5':'B' + str(sheet.max_row)]:
        for c in cell:
            address = c.value
            if address is None:
                break
            addresses.append(address)

    # Закрываем файл Excel
    workbook.close()
    return addresses


def writeXLSX(data):

    line = int(data[-1])
    workbook = openpyxl.load_workbook('balance.xlsx')
    sheet = workbook.active

    sheet.cell(row=line, column=3).value = data[0]

    result = 0
    for i in range(1, len(data)-1, 2):

        row_num = 4
        col_num = 5
        while True:
            cell_value = sheet.cell(row=row_num, column=col_num).value
            if cell_value is None:
                break

            if cell_value == data[i]:
                split_data = data[i + 1].split('$')
                value = split_data[0]
                sheet.cell(row=line, column=col_num).value = value
            col_num += 1

        if '$' in data[i+1]:
            split_data = data[i+1].split('$')
            value = split_data[-1]
            result += float(value)
            sheet.cell(row=line, column=4).value = result

    workbook.save('balance.xlsx')
    workbook.close()
